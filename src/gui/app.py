"""
app.py — Main application window for CineStats.

Builds the GUI, wires user actions to core functions, and handles all user-facing
error messages. The layout has four sections:
  1. Input   — file picker(s) + Load button
  2. Options — filter fields and output mode (changes based on report type)
  3. Output  — destination file picker + Generate Report button
  4. Status  — status bar at the bottom of the window

Processing runs on a background thread so the window stays responsive.
All text visible to the user lives in this file for easy editing.
"""

import os
import sys
import datetime
import platform
import subprocess
import threading
import traceback
import tkinter as tk
from tkinter import messagebox

# Add src/ to path so sibling modules import correctly whether run directly
# or via PyInstaller bundle.
_SRC_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _SRC_DIR not in sys.path:
    sys.path.insert(0, _SRC_DIR)

from config import REPORT_TYPE_OCCUPANCY, REPORT_TYPE_TRANSACTION
from core.reader import detect_report_type, read_file
from core.transformer import (
    filter_occupancy,
    filter_transactions,
    compute_grand_total_occupancy,
    compute_grand_total_transactions,
    summarize_occupancy_by_movie,
    summarize_transactions_by_employee,
    summarize_transactions_by_category,
)
from core.writer import write_occupancy, write_transaction_detail, write_summary
from gui.widgets import (
    FilePicker,
    LabeledEntry,
    LabeledCheckbox,
    SectionLabel,
    StatusBar,
    ReportTypeLabel,
)


# ── Layout constants ──────────────────────────────────────────────────────────
_PAD_X = 14
_PAD_Y =  6
_WIN_W = 640
_WIN_H = 560


class App:
    """
    The root application object. Owns the main window and all widgets.

    Usage:
        root = tk.Tk()
        app  = App(root)
        root.mainloop()
    """

    def __init__(self, root):
        self._root = root
        self._root.title("CineStats")
        self._root.geometry(f"{_WIN_W}x{_WIN_H}")
        self._root.resizable(True, True)
        self._root.minsize(520, 480)

        # Parsed row data cached after the user clicks "Load Files".
        # Keyed by report type; None until loaded.
        self._loaded_rows = None
        self._current_report_type = None

        self._build_ui()

    # ── UI Construction ───────────────────────────────────────────────────────

    def _build_ui(self):
        root = self._root
        root.columnconfigure(0, weight=1)

        # ── Input section ──────────────────────────────────────────────────
        input_frame = tk.LabelFrame(root, text=" Input ", padx=_PAD_X, pady=_PAD_Y)
        input_frame.grid(row=0, column=0, sticky="ew", padx=_PAD_X, pady=(_PAD_Y * 2, _PAD_Y))
        input_frame.columnconfigure(0, weight=1)

        self._file_picker = FilePicker(
            input_frame,
            label_text="File(s):",
            mode="open_multi",
        )
        self._file_picker.grid(row=0, column=0, sticky="ew", pady=(_PAD_Y, 0))

        btn_row = tk.Frame(input_frame)
        btn_row.grid(row=1, column=0, sticky="ew", pady=(_PAD_Y // 2, 0))
        btn_row.columnconfigure(1, weight=1)

        self._load_btn = tk.Button(
            btn_row,
            text="Load Files",
            command=self._on_load,
            width=12,
            bg="#2E75B6",
            fg="white",
            activebackground="#1F4E79",
            activeforeground="white",
            relief="flat",
            cursor="hand2",
        )
        self._load_btn.grid(row=0, column=0, sticky="w")

        self._report_type_label = ReportTypeLabel(btn_row)
        self._report_type_label.grid(row=0, column=1, sticky="w", padx=(12, 0))

        # Row count label — updated after loading
        self._row_count_var = tk.StringVar(value="")
        tk.Label(btn_row, textvariable=self._row_count_var, fg="#555555",
                 font=("Calibri", 9)).grid(row=0, column=2, sticky="e")

        # ── Options section ────────────────────────────────────────────────
        self._options_frame = tk.LabelFrame(root, text=" Options ", padx=_PAD_X, pady=_PAD_Y)
        self._options_frame.grid(row=1, column=0, sticky="ew", padx=_PAD_X, pady=_PAD_Y)
        self._options_frame.columnconfigure(0, weight=1)

        self._options_placeholder = tk.Label(
            self._options_frame,
            text="Load files above to see filter and output options.",
            fg="#888888",
            font=("Calibri", 9, "italic"),
        )
        self._options_placeholder.grid(row=0, column=0, sticky="w")

        # Occupancy-specific widgets (hidden until Occupancy file loaded)
        self._occ_widgets = self._build_occupancy_options(self._options_frame)
        self._hide_widgets(self._occ_widgets)

        # Transaction-specific widgets (hidden until Transaction file loaded)
        self._txn_widgets = self._build_transaction_options(self._options_frame)
        self._hide_widgets(self._txn_widgets)

        # ── Output section ─────────────────────────────────────────────────
        output_frame = tk.LabelFrame(root, text=" Output ", padx=_PAD_X, pady=_PAD_Y)
        output_frame.grid(row=2, column=0, sticky="ew", padx=_PAD_X, pady=_PAD_Y)
        output_frame.columnconfigure(0, weight=1)

        self._save_picker = FilePicker(
            output_frame,
            label_text="Save as:",
            mode="save",
        )
        self._save_picker.grid(row=0, column=0, sticky="ew", pady=(_PAD_Y, _PAD_Y // 2))

        # Checkbox: automatically open the file after export
        self._auto_open_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            output_frame,
            text="Open file after export",
            variable=self._auto_open_var,
        ).grid(row=1, column=0, sticky="w")

        # ── Generate button ────────────────────────────────────────────────
        btn_frame = tk.Frame(root)
        btn_frame.grid(row=3, column=0, pady=(_PAD_Y, _PAD_Y * 2))

        self._generate_btn = tk.Button(
            btn_frame,
            text="  Generate Report  ",
            command=self._on_generate,
            font=("Calibri", 11, "bold"),
            bg="#1F4E79",
            fg="white",
            activebackground="#2E75B6",
            activeforeground="white",
            relief="flat",
            padx=16,
            pady=8,
            cursor="hand2",
        )
        self._generate_btn.pack()

        # ── Status bar ─────────────────────────────────────────────────────
        self._status = StatusBar(root)
        self._status.grid(row=4, column=0, sticky="ew")

    def _build_occupancy_options(self, parent):
        """Creates the Occupancy filter + output-mode widgets. Returns list of widgets."""
        widgets = []

        # ── Filter sub-section ─────────────────────────────────────────────
        sep1 = SectionLabel(parent, "Filters")
        widgets.append(sep1)

        self._occ_movie = LabeledEntry(parent, "Movie contains:")
        widgets.append(self._occ_movie)

        self._occ_house = LabeledEntry(parent, "House number:", width=8)
        widgets.append(self._occ_house)

        # Date range filters — accept YYYY-MM-DD format
        date_row = tk.Frame(parent)
        widgets.append(date_row)
        date_row.columnconfigure(1, weight=1)
        date_row.columnconfigure(3, weight=1)
        tk.Label(date_row, text="Start date:", anchor="w", width=12).grid(row=0, column=0, sticky="w", padx=(0, 4))
        self._occ_start_date = tk.Entry(date_row, width=12)
        self._occ_start_date.grid(row=0, column=1, sticky="w")
        tk.Label(date_row, text="  End date:", anchor="w", width=10).grid(row=0, column=2, sticky="w", padx=(8, 4))
        self._occ_end_date = tk.Entry(date_row, width=12)
        self._occ_end_date.grid(row=0, column=3, sticky="w")
        tk.Label(date_row, text="  (YYYY-MM-DD)", fg="#888888", font=("Calibri", 8)).grid(row=0, column=4, sticky="w", padx=(4, 0))

        # ── Output mode sub-section ────────────────────────────────────────
        sep2 = SectionLabel(parent, "Output Mode")
        widgets.append(sep2)

        self._occ_output_mode = tk.StringVar(value="raw")
        modes = [
            ("Full detail — one row per showtime",     "raw"),
            ("Summary by movie — totals per film",     "by_movie"),
            ("Summary by date — totals per day",       "by_date"),
        ]
        for label, value in modes:
            rb = tk.Radiobutton(parent, text=label, variable=self._occ_output_mode, value=value, anchor="w")
            widgets.append(rb)

        return widgets

    def _build_transaction_options(self, parent):
        """Creates the Transaction filter + output-mode widgets. Returns list of widgets."""
        widgets = []

        sep1 = SectionLabel(parent, "Filters")
        widgets.append(sep1)

        self._txn_employee = LabeledEntry(parent, "Employee contains:")
        widgets.append(self._txn_employee)

        self._txn_terminal = LabeledEntry(parent, "Terminal contains:")
        widgets.append(self._txn_terminal)

        self._txn_category = LabeledEntry(parent, "Category contains:")
        widgets.append(self._txn_category)

        sep2 = SectionLabel(parent, "Output Mode")
        widgets.append(sep2)

        self._txn_output_mode = tk.StringVar(value="raw")
        modes = [
            ("Full detail — one row per item sold",            "raw"),
            ("Summary by employee — totals per staff member",  "by_employee"),
            ("Summary by category — totals per product type",  "by_category"),
        ]
        for label, value in modes:
            rb = tk.Radiobutton(parent, text=label, variable=self._txn_output_mode, value=value, anchor="w")
            widgets.append(rb)

        return widgets

    # ── Event Handlers ────────────────────────────────────────────────────────

    def _on_load(self):
        """
        Triggered by the 'Load Files' button.

        Detects the report type and parses all selected files into memory.
        Caches the result in self._loaded_rows for the Generate step.
        Runs on a background thread to keep the UI responsive during parsing.
        """
        paths = self._file_picker.get()
        if not paths:
            messagebox.showwarning("No file selected", "Please select at least one file first.")
            return

        self._load_btn.config(state="disabled")
        self._status.set_working("Reading files…")
        self._report_type_label.clear()
        self._row_count_var.set("")

        thread = threading.Thread(
            target=self._run_load,
            args=(list(paths) if isinstance(paths, list) else [paths],),
            daemon=True,
        )
        thread.start()

    def _run_load(self, paths):
        """Background worker: reads all files and caches rows."""
        try:
            # Detect type from the first file — all files must match.
            report_type = detect_report_type(paths[0])

            all_rows = []
            for path in paths:
                rtype = detect_report_type(path)
                if rtype != report_type:
                    raise ValueError(
                        f"'{os.path.basename(path)}' is a '{rtype}' report but the "
                        f"first file is '{report_type}'. All files must be the same type."
                    )
                data = read_file(path)
                all_rows.extend(data["rows"])

            if not all_rows:
                raise ValueError("No data rows found in the selected file(s).")

            self._root.after(0, self._on_load_success, report_type, all_rows)

        except Exception as exc:
            tb = traceback.format_exc()
            self._root.after(0, self._on_load_error, str(exc), tb)

    def _on_load_success(self, report_type, rows):
        """Called on the main thread after successful file load."""
        self._loaded_rows = rows
        self._current_report_type = report_type
        self._load_btn.config(state="normal")
        self._report_type_label.set(report_type)
        self._row_count_var.set(f"{len(rows):,} rows loaded")
        self._show_filters_for(report_type)
        self._status.set_success(f"Loaded {len(rows):,} rows from {len(self._file_picker.get())} file(s).")

    def _on_load_error(self, message, tb):
        """Called on the main thread after a failed load."""
        self._load_btn.config(state="normal")
        self._loaded_rows = None
        self._current_report_type = None
        self._status.set_error(message)
        messagebox.showerror("Error loading file", message)
        print("=== CineStats load error ===")
        print(tb)

    def _on_generate(self):
        """
        Triggered by the 'Generate Report' button.

        Validates inputs, then runs the export on a background thread.
        """
        if self._loaded_rows is None:
            messagebox.showwarning("No data loaded", "Please select and load files first.")
            return

        save_path = self._save_picker.get()
        if not save_path:
            messagebox.showwarning("No output file", "Please choose where to save the output.")
            return

        self._generate_btn.config(state="disabled")
        self._status.set_working("Generating report…")

        thread = threading.Thread(
            target=self._run_export,
            args=(
                self._loaded_rows,
                save_path,
                self._current_report_type,
                self._collect_options(),
                self._auto_open_var.get(),
            ),
            daemon=True,
        )
        thread.start()

    def _run_export(self, rows, save_path, report_type, options, auto_open):
        """Background worker: applies filters, writes xlsx."""
        try:
            if report_type == REPORT_TYPE_OCCUPANCY:
                output_mode = options.get("output_mode", "raw")
                filtered = filter_occupancy(
                    rows,
                    start_date=options.get("start_date"),
                    end_date=options.get("end_date"),
                    movie=options.get("movie") or None,
                    house=int(options["house"]) if options.get("house") else None,
                )
                row_count = self._write_occupancy_output(filtered, save_path, output_mode)

            elif report_type == REPORT_TYPE_TRANSACTION:
                output_mode = options.get("output_mode", "raw")
                filtered = filter_transactions(
                    rows,
                    employee=options.get("employee") or None,
                    terminal=options.get("terminal") or None,
                    category=options.get("category") or None,
                )
                row_count = self._write_transaction_output(filtered, save_path, output_mode)

            else:
                raise ValueError(f"Unknown report type: {report_type!r}")

            self._root.after(0, self._on_export_success, save_path, row_count, auto_open)

        except Exception as exc:
            tb = traceback.format_exc()
            self._root.after(0, self._on_export_error, str(exc), tb)

    def _write_occupancy_output(self, filtered_rows, save_path, output_mode):
        """
        Writes occupancy data to xlsx using the chosen output mode.

        Returns the number of data rows written (excluding the header/total rows).
        """
        if output_mode == "raw":
            grand_total = compute_grand_total_occupancy(filtered_rows)
            write_occupancy(filtered_rows, save_path, grand_total=grand_total)
            return len(filtered_rows)

        elif output_mode == "by_movie":
            summary_rows = summarize_occupancy_by_movie(filtered_rows)
            columns = ["movie", "seats_sold", "total_seats", "occupancy_pct", "box_gross"]
            headers = ["Movie", "Seats Sold", "Total Seats", "Occupancy %", "Box Gross ($)"]
            # Build grand total row for the summary
            gt = compute_grand_total_occupancy(filtered_rows)
            total_row = ["TOTAL", gt["seats_sold"], gt["total_seats"], gt["occupancy_pct"], gt["box_gross"]]
            # Write manually using the generic writer
            _write_summary_with_headers(summary_rows, columns, headers, "By Movie", save_path, total_row)
            return len(summary_rows)

        elif output_mode == "by_date":
            from core.transformer import summarize_occupancy_by_date
            summary_rows = summarize_occupancy_by_date(filtered_rows)
            columns = ["date", "seats_sold", "total_seats", "occupancy_pct", "box_gross"]
            headers = ["Date", "Seats Sold", "Total Seats", "Occupancy %", "Box Gross ($)"]
            gt = compute_grand_total_occupancy(filtered_rows)
            total_row = ["TOTAL", gt["seats_sold"], gt["total_seats"], gt["occupancy_pct"], gt["box_gross"]]
            _write_summary_with_headers(summary_rows, columns, headers, "By Date", save_path, total_row)
            return len(summary_rows)

        raise ValueError(f"Unknown output mode: {output_mode!r}")

    def _write_transaction_output(self, filtered_rows, save_path, output_mode):
        """
        Writes transaction data to xlsx using the chosen output mode.

        Returns the number of data rows written.
        """
        if output_mode == "raw":
            grand_total = compute_grand_total_transactions(filtered_rows)
            write_transaction_detail(filtered_rows, save_path, grand_total=grand_total)
            return len(filtered_rows)

        elif output_mode == "by_employee":
            summary_rows = summarize_transactions_by_employee(filtered_rows)
            columns = ["employee", "transaction_count", "item_count", "total_sales"]
            headers = ["Employee", "Transaction Count", "Item Count", "Total Sales ($)"]
            gt = compute_grand_total_transactions(filtered_rows)
            total_row = ["TOTAL", gt["transaction_count"], gt["item_count"], gt["total_sales"]]
            _write_summary_with_headers(summary_rows, columns, headers, "By Employee", save_path, total_row)
            return len(summary_rows)

        elif output_mode == "by_category":
            summary_rows = summarize_transactions_by_category(filtered_rows)
            columns = ["category", "item_count", "total_quantity", "total_sales"]
            headers = ["Category", "Item Count", "Total Quantity", "Total Sales ($)"]
            total_row = [
                "TOTAL",
                sum(r["item_count"] for r in summary_rows),
                sum(r["total_quantity"] for r in summary_rows),
                sum(r["total_sales"] for r in summary_rows),
            ]
            _write_summary_with_headers(summary_rows, columns, headers, "By Category", save_path, total_row)
            return len(summary_rows)

        raise ValueError(f"Unknown output mode: {output_mode!r}")

    def _on_export_success(self, save_path, row_count, auto_open):
        """Called on the main thread after a successful export."""
        self._generate_btn.config(state="normal")
        filename = os.path.basename(save_path)
        self._status.set_success(f"Saved {row_count:,} rows → {filename}")
        messagebox.showinfo(
            "Export complete",
            f"Report saved successfully:\n{save_path}\n\n{row_count:,} data rows exported.",
        )
        if auto_open:
            _open_file(save_path)

    def _on_export_error(self, message, tb):
        """Called on the main thread after a failed export."""
        self._generate_btn.config(state="normal")
        self._status.set_error(message)
        messagebox.showerror("Export failed", message)
        print("=== CineStats export error ===")
        print(tb)

    # ── Filter / options visibility ───────────────────────────────────────────

    def _show_filters_for(self, report_type):
        """Shows the correct options section and hides the other."""
        self._options_placeholder.grid_remove()

        if report_type == REPORT_TYPE_OCCUPANCY:
            self._hide_widgets(self._txn_widgets)
            self._show_widgets(self._occ_widgets)
        elif report_type == REPORT_TYPE_TRANSACTION:
            self._hide_widgets(self._occ_widgets)
            self._show_widgets(self._txn_widgets)

    def _show_widgets(self, widget_list):
        for i, widget in enumerate(widget_list):
            widget.grid(row=i, column=0, sticky="ew", pady=2)

    def _hide_widgets(self, widget_list):
        for widget in widget_list:
            widget.grid_remove()

    # ── Collect options ───────────────────────────────────────────────────────

    def _collect_options(self):
        """
        Reads current filter values and output mode from the visible widgets.

        Returns a dict whose keys depend on the current report type.
        """
        if self._current_report_type == REPORT_TYPE_OCCUPANCY:
            return {
                "movie":       self._occ_movie.get(),
                "house":       self._occ_house.get(),
                "start_date":  _parse_date_entry(self._occ_start_date.get()),
                "end_date":    _parse_date_entry(self._occ_end_date.get()),
                "output_mode": self._occ_output_mode.get(),
            }
        elif self._current_report_type == REPORT_TYPE_TRANSACTION:
            return {
                "employee":    self._txn_employee.get(),
                "terminal":    self._txn_terminal.get(),
                "category":    self._txn_category.get(),
                "output_mode": self._txn_output_mode.get(),
            }
        return {}


# ── Module-level helpers ──────────────────────────────────────────────────────


def _parse_date_entry(text):
    """
    Converts a date string typed by the user into a datetime.date object.

    Accepts YYYY-MM-DD format. Returns None if the field is empty or unparseable,
    so the caller can treat None as "no date filter".
    """
    text = text.strip()
    if not text:
        return None
    try:
        return datetime.datetime.strptime(text, "%Y-%m-%d").date()
    except ValueError:
        # Surface the error as a messagebox from the caller context.
        raise ValueError(
            f"Could not parse date '{text}'.\nPlease use the format YYYY-MM-DD (e.g. 2026-04-01)."
        )


def _open_file(filepath):
    """
    Opens a file with the OS default application.

    Works on Windows (os.startfile), macOS (open), and Linux (xdg-open).
    Fails silently if the OS doesn't support it — never crashes the app.
    """
    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(filepath)                        # Built-in Windows-only function
        elif system == "Darwin":
            subprocess.Popen(["open", filepath])
        else:
            # Linux: xdg-open is available on most desktop environments
            subprocess.Popen(["xdg-open", filepath])
    except Exception:
        pass  # Auto-open failing is non-critical; the file was still saved successfully


def _write_summary_with_headers(rows, column_keys, headers, sheet_title, filepath, total_row=None):
    """
    Writes summary row dicts to an xlsx file.

    Args:
        rows:        list of dicts
        column_keys: list of dict keys to extract, in column order
        headers:     list of human-readable header strings (same length as column_keys)
        sheet_title: name for the worksheet tab
        filepath:    destination path
        total_row:   optional list of values for a grand-total row
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    import datetime as dt

    _HEADER_BG  = "1F4E79"
    _HEADER_FG  = "FFFFFF"
    _ROW_ALT    = "D6E4F0"
    _TOTAL_BG   = "2E75B6"
    _TOTAL_FG   = "FFFFFF"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Header row
    hfont = Font(bold=True, color=_HEADER_FG, name="Calibri", size=11)
    hfill = PatternFill(fill_type="solid", fgColor=_HEADER_BG)
    for col_i, label in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.font = hfont
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # Data rows
    dfont = Font(name="Calibri", size=10)
    afill = PatternFill(fill_type="solid", fgColor=_ROW_ALT)
    for row_i, row_dict in enumerate(rows, 2):
        fill = afill if row_i % 2 == 1 else None
        for col_i, key in enumerate(column_keys, 1):
            value = row_dict.get(key, "")
            if isinstance(value, float):
                value = round(value, 2)
            cell = ws.cell(row=row_i, column=col_i, value=value)
            cell.font = dfont
            if fill:
                cell.fill = fill

    # Grand total row
    if total_row is not None:
        tr = len(rows) + 2
        tfont = Font(bold=True, color=_TOTAL_FG, name="Calibri", size=11)
        tfill = PatternFill(fill_type="solid", fgColor=_TOTAL_BG)
        for col_i, value in enumerate(total_row, 1):
            if isinstance(value, float):
                value = round(value, 2)
            cell = ws.cell(row=tr, column=col_i, value=value)
            cell.font = tfont
            cell.fill = tfill

    # Auto-size columns
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

    wb.save(filepath)
