"""
reader.py — Reads Cinemark-exported .xlsx files and returns clean Python data structures.

Each public function returns a dict with two keys:
  "meta"  — report-level info (theater number, date, etc.)
  "rows"  — list of dicts, one dict per data record

What this file does NOT do: filter, sort, aggregate, or write anything.
All transformation belongs in transformer.py.
"""

import re
import datetime

import openpyxl

from config import (
    REPORT_TYPE_OCCUPANCY,
    REPORT_TYPE_TRANSACTION,
    OCC_DATA_START_ROW,
    OCC_COL_MOVIE,
    OCC_COL_HOUSE,
    OCC_COL_SHOWTIME,
    OCC_COL_SEATS_SOLD,
    OCC_COL_TOTAL_SEATS,
    OCC_COL_OCCUPANCY_PCT,
    OCC_COL_BOX_GROSS,
    OCC_MOVIE_TOTALS_SENTINEL,
    OCC_DELETED_MARKER,
    TXN_DATA_START_ROW,
    TXN_COL_TIME,
    TXN_COL_TOTAL,
    TXN_COL_SALE_TYPE,
    TXN_COL_TXN_ID,
    TXN_COL_TERMINAL,
    TXN_COL_EMPLOYEE,
    TXN_COL_ITEM_TYPE,
    TXN_COL_CATEGORY,
    TXN_COL_QUANTITY,
    TXN_COL_UNIT_PRICE,
    TXN_COL_PRETAX,
    TXN_COL_TAX,
    TXN_COL_POSTTAX,
    TXN_HEADER_SENTINEL_TIME,
    TXN_HEADER_SENTINEL_TOTAL,
    TXN_ITEM_SENTINEL,
    TXN_PAYMENT_SENTINEL,
    TXN_ITEM_HEADER_SENTINEL,
)

# Pattern that matches a Cinemark date separator row like "04/26/2026".
_DATE_PATTERN = re.compile(r"^\d{2}/\d{2}/\d{4}$")


# ── Public API ────────────────────────────────────────────────────────────────


def detect_report_type(filepath):
    """
    Opens the xlsx file at filepath and reads cell A1 to determine the report type.

    Returns:
        REPORT_TYPE_OCCUPANCY or REPORT_TYPE_TRANSACTION (strings from config.py)

    Raises:
        ValueError: if the file does not look like a known Cinemark report.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        title_cell = ws.cell(row=1, column=1).value
    finally:
        wb.close()

    if title_cell == REPORT_TYPE_OCCUPANCY:
        return REPORT_TYPE_OCCUPANCY
    if title_cell == REPORT_TYPE_TRANSACTION:
        return REPORT_TYPE_TRANSACTION

    raise ValueError(
        f"Unrecognised report type in '{filepath}'.\n"
        f"Cell A1 contains: {title_cell!r}\n"
        f"Expected: {REPORT_TYPE_OCCUPANCY!r} or {REPORT_TYPE_TRANSACTION!r}"
    )


def read_file(filepath):
    """
    Auto-detects the report type and parses the file.

    Returns:
        dict with keys:
            "report_type" — REPORT_TYPE_OCCUPANCY or REPORT_TYPE_TRANSACTION
            "meta"        — dict of report-level metadata
            "rows"        — list of row dicts (one per data record)
    """
    report_type = detect_report_type(filepath)

    if report_type == REPORT_TYPE_OCCUPANCY:
        result = read_occupancy(filepath)
    else:
        result = read_transaction_detail(filepath)

    result["report_type"] = report_type
    return result


def read_occupancy(filepath):
    """
    Parses a Cinemark Occupancy .xlsx report.

    The raw Occupancy report has a nested structure:
      - Date separator rows (grouping showtimes by calendar date)
      - Showtime rows (one per screening, no movie name)
      - Movie Totals rows (the movie name appears HERE, at the bottom of its group)

    This function reconstructs which movie each showtime belongs to by reading the
    Movie Totals row that follows each batch of showtime rows.

    Returns:
        dict with keys:
            "meta" — {"theater": int, "start_date": datetime, "end_date": datetime}
            "rows" — list of dicts, each with keys matching OCC_OUTPUT_COLUMNS
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        meta = _parse_occupancy_meta(ws)
        rows = _parse_occupancy_rows(ws, meta)
    finally:
        wb.close()

    return {"meta": meta, "rows": rows}


def read_transaction_detail(filepath):
    """
    Parses a Cinemark Transaction Detail .xlsx report.

    The raw Transaction Detail report has a repeated block structure per transaction:
      - Transaction header label row  (contains "Time", "Total" labels)
      - Transaction value row         (actual time, total, employee, terminal, ID)
      - "Transaction Detail:" separator row
      - Item column-header row
      - One or more item rows         (product sold, quantity, price, tax)
      - Optional ticket-detail rows   (multiline seat/ticket info — skipped)
      - "Payment Detail:" section     (payment method — skipped)

    This function flattens each transaction into one row per item, copying the
    transaction-level fields (employee, terminal, etc.) onto every item row.

    Returns:
        dict with keys:
            "meta" — {"theater": int, "date": datetime}
            "rows" — list of dicts, each with keys matching TXN_OUTPUT_COLUMNS
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        ws = wb.active
        meta = _parse_transaction_meta(ws)
        rows = _parse_transaction_rows(ws, meta)
    finally:
        wb.close()

    return {"meta": meta, "rows": rows}


# ── Internal helpers ──────────────────────────────────────────────────────────


def _parse_occupancy_meta(ws):
    """Reads theater number and date range from the Occupancy report header rows."""
    # Row 3: [None, 'Start Date: ', None, <datetime>, None, None, None, 'End Date: ', None, <datetime>]
    date_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
    # Row 4: [None, None, 'Theater: ', <int>]
    theater_row = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))[0]

    return {
        "start_date": date_row[3],    # datetime object from Excel
        "end_date":   date_row[9],    # datetime object from Excel
        "theater":    theater_row[3], # integer theater number
    }


def _parse_occupancy_rows(ws, meta):
    """
    Iterates data rows and reconstructs showtime records with movie names.

    Strategy: showtime rows appear before their movie name is known. The movie name
    only appears on the "Movie Totals" row at the end of each movie's group. So we
    buffer incoming showtime rows and flush them once we see the Movie Totals row.

    Args:
        ws:   openpyxl worksheet (read-only)
        meta: dict from _parse_occupancy_meta

    Returns:
        list of showtime-record dicts
    """
    result = []
    current_date = None

    # Buffer for showtime rows that are waiting for their movie name to appear.
    # Flushed to `result` when the corresponding Movie Totals row is encountered.
    pending_showtimes = []

    for row in ws.iter_rows(min_row=OCC_DATA_START_ROW, values_only=True):
        # Skip rows that are entirely empty (blank spacer rows in the report).
        if not any(v is not None for v in row):
            continue

        col_movie    = row[OCC_COL_MOVIE]
        col_house    = row[OCC_COL_HOUSE]
        col_showtime = row[OCC_COL_SHOWTIME]
        col_sold     = row[OCC_COL_SEATS_SOLD]
        col_total    = row[OCC_COL_TOTAL_SEATS]
        col_pct      = row[OCC_COL_OCCUPANCY_PCT]
        col_gross    = row[OCC_COL_BOX_GROSS]

        # ── Date separator row ─────────────────────────────────────────────
        # Cinemark groups showtimes under date rows formatted as "MM/DD/YYYY".
        # These rows have no house, showtime, or seat data.
        if isinstance(col_movie, str) and _DATE_PATTERN.match(col_movie):
            current_date = _parse_date_string(col_movie)
            continue

        # ── Movie Totals row ───────────────────────────────────────────────
        # This row carries the movie name and appears after all of its showtimes.
        # We use it to back-fill the movie name on all buffered showtime rows.
        if col_showtime == OCC_MOVIE_TOTALS_SENTINEL:
            movie_name = col_movie or "Unknown"
            for record in pending_showtimes:
                record["movie"] = movie_name
                result.append(record)
            pending_showtimes = []
            continue

        # ── Showtime row ───────────────────────────────────────────────────
        # A showtime row always has a numeric house number.
        if col_house is not None and isinstance(col_house, (int, float)):
            showtime_str = str(col_showtime) if col_showtime else ""

            # Skip showtimes Cinemark has marked as deleted/cancelled.
            # These still appear in the export but should not count toward totals.
            if OCC_DELETED_MARKER in showtime_str:
                continue

            pending_showtimes.append({
                "date":          current_date,
                "movie":         None,  # filled in when Movie Totals row is found
                "house":         int(col_house),
                "showtime":      showtime_str,
                "seats_sold":    int(col_sold) if col_sold is not None else 0,
                "total_seats":   int(col_total) if col_total is not None else 0,
                "occupancy_pct": col_pct or "0.00%",
                "box_gross":     _parse_currency(col_gross),
            })

    # Any remaining buffered rows (shouldn't happen in well-formed files, but guard anyway).
    for record in pending_showtimes:
        if record["movie"] is None:
            record["movie"] = "Unknown"
        result.append(record)

    return result


def _parse_transaction_meta(ws):
    """Reads theater number and date from the Transaction Detail report header rows."""
    # Row 2: [None, 'Date: ', None, None, None, <datetime>]
    date_row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
    # Row 3: ['Theater: ', None, None, None, None, <int>]
    theater_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    return {
        "date":    date_row[5],      # datetime object from Excel
        "theater": theater_row[5],   # integer theater number
    }


def _parse_transaction_rows(ws, meta):
    """
    Iterates data rows and flattens transaction + item data into one row per item.

    Strategy: when we see a transaction-value row (identified by a numeric Transaction ID),
    we store its fields in `current_txn`. When we then see an item row (identified by a
    numeric Quantity field), we combine it with `current_txn` and emit one output record.

    Args:
        ws:   openpyxl worksheet (read-only)
        meta: dict from _parse_transaction_meta

    Returns:
        list of flattened item-record dicts
    """
    result = []

    # Holds the transaction-level fields for the transaction currently being parsed.
    # Reset each time we encounter a new transaction-value row.
    current_txn = {}

    for row in ws.iter_rows(min_row=TXN_DATA_START_ROW, values_only=True):
        # Skip entirely empty rows.
        if not any(v is not None for v in row):
            continue

        col_d = row[TXN_COL_TIME]       # Time string or label sentinel
        col_e = row[TXN_COL_TOTAL]      # Total dollar amount or label sentinel

        # ── Transaction header LABEL row ───────────────────────────────────
        # This row contains the column labels ("Time", "Total", "Transaction ID"…)
        # and is not data — skip it.
        if col_d == TXN_HEADER_SENTINEL_TIME and col_e == TXN_HEADER_SENTINEL_TOTAL:
            continue

        # ── Separator / section-header rows ────────────────────────────────
        # "Transaction Detail:" and "Payment Detail:" rows are structural markers.
        # The Item column-header row ("Item", "Product Name"…) is also skipped.
        if col_d in (TXN_ITEM_SENTINEL, TXN_PAYMENT_SENTINEL, TXN_ITEM_HEADER_SENTINEL):
            continue

        # ── Transaction value row ──────────────────────────────────────────
        # Identified by: col D is a time string AND col M is an integer (Transaction ID).
        # The sale-type column (col I = " - SALE") additionally confirms this is a sale.
        txn_id = row[TXN_COL_TXN_ID]
        if isinstance(txn_id, int) and isinstance(col_d, str) and ":" in col_d:
            current_txn = {
                "date":        meta["date"],
                "time":        col_d,
                "txn_id":      txn_id,
                "terminal":    str(row[TXN_COL_TERMINAL] or "").strip(),
                "employee":    str(row[TXN_COL_EMPLOYEE] or "").strip(),
                "txn_total":   _parse_currency(row[TXN_COL_TOTAL]),
            }
            continue

        # ── Item row ───────────────────────────────────────────────────────
        # Identified by: col N (Quantity) is a float/int.
        # Ticket-detail rows (multiline seat info) do NOT have a quantity, so this
        # check correctly excludes them.
        quantity = row[TXN_COL_QUANTITY]
        if isinstance(quantity, (int, float)) and current_txn:
            result.append({
                **current_txn,
                "item_type":    str(col_d or "").strip(),
                "category":     str(row[TXN_COL_CATEGORY] or "").strip(),
                "quantity":     float(quantity),
                "unit_price":   _parse_currency(row[TXN_COL_UNIT_PRICE]),
                "pretax":       _parse_currency(row[TXN_COL_PRETAX]),
                "tax":          _parse_currency_or_dash(row[TXN_COL_TAX]),
                "posttax":      _parse_currency(row[TXN_COL_POSTTAX]),
            })

    return result


# ── Value-parsing utilities ───────────────────────────────────────────────────


def _parse_currency(value):
    """
    Converts a Cinemark currency string to a float.

    Handles: "$20.00" → 20.0, "$1,149.50" → 1149.5, "$0.00" → 0.0,
             "(9.50)" → -9.5  (accounting-style negatives),
             "" → 0.0, None → 0.0, 20.0 → 20.0 (already numeric).
    """
    if value is None or value == "":
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = str(value).replace("$", "").replace(",", "").strip()
    if not cleaned:
        return 0.0
    # Accounting notation wraps negatives in parentheses: (9.50) → -9.50
    if cleaned.startswith("(") and cleaned.endswith(")"):
        return -float(cleaned[1:-1])
    return float(cleaned)


def _parse_currency_or_dash(value):
    """
    Like _parse_currency but also handles the "-" sentinel Cinemark uses for zero tax.

    Cinemark writes "-" in the Tax column when no tax applies (e.g. movie tickets).
    """
    if value == "-":
        return 0.0
    return _parse_currency(value)


def _parse_date_string(date_str):
    """
    Converts a Cinemark date string "MM/DD/YYYY" to a Python date object.

    Returns None if the string cannot be parsed, so callers get a safe fallback
    rather than an unhandled exception on a malformed row.
    """
    try:
        return datetime.datetime.strptime(date_str, "%m/%d/%Y").date()
    except (ValueError, TypeError):
        return None
