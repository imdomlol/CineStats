"""
writer.py — Writes processed CineStats data to a formatted .xlsx file.

All output styling is applied here. This file handles:
  - Bold header row
  - Auto-sized column widths
  - Correct cell types (dates as dates, numbers as numbers, strings as strings)
  - Optional grand-total row at the bottom (bold, highlighted)
  - Alternating row shading for readability

What this file does NOT do: parse input, filter data, or interact with the GUI.
"""

import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import (
    REPORT_TYPE_OCCUPANCY,
    REPORT_TYPE_TRANSACTION,
    OCC_OUTPUT_COLUMNS,
    TXN_OUTPUT_COLUMNS,
)


# ── Colour constants ──────────────────────────────────────────────────────────
# Using subtle, professional colours that print well in grey-scale too.

_COLOUR_HEADER_BG  = "1F4E79"  # Dark navy for the header row background
_COLOUR_HEADER_FG  = "FFFFFF"  # White text on dark header
_COLOUR_ROW_ALT    = "D6E4F0"  # Soft blue for alternating data rows
_COLOUR_TOTAL_BG   = "2E75B6"  # Medium navy for the grand-total row
_COLOUR_TOTAL_FG   = "FFFFFF"  # White text on total row

# Number formats applied to specific column types.
_FMT_CURRENCY = '#,##0.00'     # e.g. 1234.50  →  1,234.50
_FMT_DATE     = 'YYYY-MM-DD'   # e.g. 2026-04-26


# ── Public API ────────────────────────────────────────────────────────────────


def write_occupancy(rows, filepath, grand_total=None):
    """
    Writes a list of occupancy row dicts to a formatted .xlsx file.

    Args:
        rows:        list of occupancy row dicts (from reader + transformer)
        filepath:    destination path string (will be created or overwritten)
        grand_total: optional dict from transformer.compute_grand_total_occupancy()
                     — if provided, a summary row is appended at the bottom

    Raises:
        IOError: if the file cannot be written (e.g. it is open in Excel)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Occupancy"

    # Write the column header row.
    _write_header(ws, OCC_OUTPUT_COLUMNS)

    # Write one data row per showtime record.
    for i, row in enumerate(rows, start=2):
        data = [
            row.get("date"),           # date object  → formatted as date
            row.get("movie", ""),
            row.get("house", ""),
            row.get("showtime", ""),
            row.get("seats_sold", 0),
            row.get("total_seats", 0),
            row.get("occupancy_pct", "0.00%"),
            row.get("box_gross", 0.0), # float → formatted as currency
        ]
        _write_data_row(ws, i, data, alternate=(i % 2 == 1))

    # Apply currency format to the Box Gross column (column 8).
    _apply_column_format(ws, col_index=8, fmt=_FMT_CURRENCY, start_row=2)

    # Apply date format to the Date column (column 1).
    _apply_column_format(ws, col_index=1, fmt=_FMT_DATE, start_row=2)

    # Append grand-total row if provided.
    if grand_total is not None:
        total_row_num = len(rows) + 2
        data = [
            "TOTAL", "", "", "",
            grand_total.get("seats_sold", 0),
            grand_total.get("total_seats", 0),
            grand_total.get("occupancy_pct", "0.00%"),
            grand_total.get("box_gross", 0.0),
        ]
        _write_total_row(ws, total_row_num, data)
        _apply_column_format(ws, col_index=8, fmt=_FMT_CURRENCY, start_row=total_row_num, end_row=total_row_num)

    _auto_size_columns(ws)
    wb.save(filepath)


def write_transaction_detail(rows, filepath, grand_total=None):
    """
    Writes a list of flattened transaction item dicts to a formatted .xlsx file.

    Args:
        rows:        list of transaction item dicts (from reader + transformer)
        filepath:    destination path string (will be created or overwritten)
        grand_total: optional dict from transformer.compute_grand_total_transactions()

    Raises:
        IOError: if the file cannot be written
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction Detail"

    _write_header(ws, TXN_OUTPUT_COLUMNS)

    for i, row in enumerate(rows, start=2):
        data = [
            row.get("date"),           # date object
            row.get("time", ""),
            row.get("txn_id", ""),
            row.get("terminal", ""),
            row.get("employee", ""),
            row.get("item_type", ""),
            row.get("category", ""),
            row.get("quantity", 0.0),
            row.get("unit_price", 0.0),
            row.get("pretax", 0.0),
            row.get("tax", 0.0),
            row.get("posttax", 0.0),
            row.get("txn_total", 0.0),
        ]
        _write_data_row(ws, i, data, alternate=(i % 2 == 1))

    # Apply currency format to all dollar columns (unit_price, pretax, tax, posttax, txn_total).
    for col in [9, 10, 11, 12, 13]:
        _apply_column_format(ws, col_index=col, fmt=_FMT_CURRENCY, start_row=2)

    _apply_column_format(ws, col_index=1, fmt=_FMT_DATE, start_row=2)

    if grand_total is not None:
        total_row_num = len(rows) + 2
        # Spread the summary across the same columns as the data.
        data = [
            "TOTAL", "", "", "", "", "", "",
            "",  # quantity — not meaningful as a total
            "",  # unit price
            "",  # pretax
            "",  # tax
            "",  # posttax per item
            grand_total.get("total_sales", 0.0),
        ]
        _write_total_row(ws, total_row_num, data)
        # Also annotate transaction and item counts in the notes column.
        ws.cell(row=total_row_num, column=3).value = (
            f"{grand_total.get('transaction_count', 0)} transactions, "
            f"{grand_total.get('item_count', 0)} items"
        )
        _apply_column_format(ws, col_index=13, fmt=_FMT_CURRENCY, start_row=total_row_num, end_row=total_row_num)

    _auto_size_columns(ws)
    wb.save(filepath)


def write_occupancy_full(rows, filepath, grand_total, by_movie_rows, by_date_rows):
    """
    Writes a multi-sheet Occupancy workbook containing:
      Sheet 1: "Detail"    — one row per showtime (same as write_occupancy)
      Sheet 2: "By Movie"  — totals per film
      Sheet 3: "By Date"   — totals per calendar day

    This is the "Complete Report" output mode — everything in one file.

    Args:
        rows:          list of occupancy row dicts (filtered)
        filepath:      destination path
        grand_total:   dict from compute_grand_total_occupancy()
        by_movie_rows: list from summarize_occupancy_by_movie()
        by_date_rows:  list from summarize_occupancy_by_date()
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Full detail ───────────────────────────────────────────────
    ws_detail = wb.active
    ws_detail.title = "Detail"
    _write_header(ws_detail, OCC_OUTPUT_COLUMNS)
    for i, row in enumerate(rows, start=2):
        data = [
            row.get("date"),
            row.get("movie", ""),
            row.get("house", ""),
            row.get("showtime", ""),
            row.get("seats_sold", 0),
            row.get("total_seats", 0),
            row.get("occupancy_pct", "0.00%"),
            row.get("box_gross", 0.0),
        ]
        _write_data_row(ws_detail, i, data, alternate=(i % 2 == 1))
    _apply_column_format(ws_detail, col_index=8, fmt=_FMT_CURRENCY, start_row=2)
    _apply_column_format(ws_detail, col_index=1, fmt=_FMT_DATE, start_row=2)
    total_r = len(rows) + 2
    _write_total_row(ws_detail, total_r, [
        "TOTAL", "", "", "",
        grand_total["seats_sold"], grand_total["total_seats"],
        grand_total["occupancy_pct"], grand_total["box_gross"],
    ])
    _apply_column_format(ws_detail, col_index=8, fmt=_FMT_CURRENCY, start_row=total_r, end_row=total_r)
    _auto_size_columns(ws_detail)

    # ── Sheet 2: By Movie ──────────────────────────────────────────────────
    ws_movie = wb.create_sheet("By Movie")
    movie_headers = ["Movie", "Seats Sold", "Total Seats", "Occupancy %", "Box Gross ($)"]
    movie_keys    = ["movie", "seats_sold", "total_seats", "occupancy_pct", "box_gross"]
    _write_header(ws_movie, movie_headers)
    for i, row in enumerate(by_movie_rows, start=2):
        _write_data_row(ws_movie, i, [row.get(k, "") for k in movie_keys], alternate=(i % 2 == 1))
    _apply_column_format(ws_movie, col_index=5, fmt=_FMT_CURRENCY, start_row=2)
    _write_total_row(ws_movie, len(by_movie_rows) + 2, [
        "TOTAL",
        grand_total["seats_sold"], grand_total["total_seats"],
        grand_total["occupancy_pct"], grand_total["box_gross"],
    ])
    _auto_size_columns(ws_movie)

    # ── Sheet 3: By Date ───────────────────────────────────────────────────
    ws_date = wb.create_sheet("By Date")
    date_headers = ["Date", "Seats Sold", "Total Seats", "Occupancy %", "Box Gross ($)"]
    date_keys    = ["date", "seats_sold", "total_seats", "occupancy_pct", "box_gross"]
    _write_header(ws_date, date_headers)
    for i, row in enumerate(by_date_rows, start=2):
        _write_data_row(ws_date, i, [row.get(k, "") for k in date_keys], alternate=(i % 2 == 1))
    _apply_column_format(ws_date, col_index=5, fmt=_FMT_CURRENCY, start_row=2)
    _apply_column_format(ws_date, col_index=1, fmt=_FMT_DATE, start_row=2)
    _write_total_row(ws_date, len(by_date_rows) + 2, [
        "TOTAL",
        grand_total["seats_sold"], grand_total["total_seats"],
        grand_total["occupancy_pct"], grand_total["box_gross"],
    ])
    _auto_size_columns(ws_date)

    wb.save(filepath)


def write_transaction_full(rows, filepath, grand_total, by_employee_rows, by_category_rows):
    """
    Writes a multi-sheet Transaction Detail workbook containing:
      Sheet 1: "Detail"      — one row per item sold
      Sheet 2: "By Employee" — totals per staff member
      Sheet 3: "By Category" — totals per product category

    Args:
        rows:               list of transaction item dicts (filtered)
        filepath:           destination path
        grand_total:        dict from compute_grand_total_transactions()
        by_employee_rows:   list from summarize_transactions_by_employee()
        by_category_rows:   list from summarize_transactions_by_category()
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Full detail ───────────────────────────────────────────────
    ws_detail = wb.active
    ws_detail.title = "Detail"
    _write_header(ws_detail, TXN_OUTPUT_COLUMNS)
    for i, row in enumerate(rows, start=2):
        data = [
            row.get("date"),
            row.get("time", ""),
            row.get("txn_id", ""),
            row.get("terminal", ""),
            row.get("employee", ""),
            row.get("item_type", ""),
            row.get("category", ""),
            row.get("quantity", 0.0),
            row.get("unit_price", 0.0),
            row.get("pretax", 0.0),
            row.get("tax", 0.0),
            row.get("posttax", 0.0),
            row.get("txn_total", 0.0),
        ]
        _write_data_row(ws_detail, i, data, alternate=(i % 2 == 1))
    for col in [9, 10, 11, 12, 13]:
        _apply_column_format(ws_detail, col_index=col, fmt=_FMT_CURRENCY, start_row=2)
    _apply_column_format(ws_detail, col_index=1, fmt=_FMT_DATE, start_row=2)
    total_r = len(rows) + 2
    total_data = ["TOTAL", "", "", "", "", "", "", "", "", "", "", "", grand_total["total_sales"]]
    _write_total_row(ws_detail, total_r, total_data)
    ws_detail.cell(row=total_r, column=3).value = (
        f"{grand_total['transaction_count']} transactions, {grand_total['item_count']} items"
    )
    _apply_column_format(ws_detail, col_index=13, fmt=_FMT_CURRENCY, start_row=total_r, end_row=total_r)
    _auto_size_columns(ws_detail)

    # ── Sheet 2: By Employee ───────────────────────────────────────────────
    ws_emp = wb.create_sheet("By Employee")
    emp_headers = ["Employee", "Transaction Count", "Item Count", "Total Sales ($)"]
    emp_keys    = ["employee", "transaction_count", "item_count", "total_sales"]
    _write_header(ws_emp, emp_headers)
    for i, row in enumerate(by_employee_rows, start=2):
        _write_data_row(ws_emp, i, [row.get(k, "") for k in emp_keys], alternate=(i % 2 == 1))
    _apply_column_format(ws_emp, col_index=4, fmt=_FMT_CURRENCY, start_row=2)
    _write_total_row(ws_emp, len(by_employee_rows) + 2, [
        "TOTAL",
        grand_total["transaction_count"],
        grand_total["item_count"],
        grand_total["total_sales"],
    ])
    _auto_size_columns(ws_emp)

    # ── Sheet 3: By Category ───────────────────────────────────────────────
    ws_cat = wb.create_sheet("By Category")
    cat_headers = ["Category", "Item Count", "Total Quantity", "Total Sales ($)"]
    cat_keys    = ["category", "item_count", "total_quantity", "total_sales"]
    _write_header(ws_cat, cat_headers)
    for i, row in enumerate(by_category_rows, start=2):
        _write_data_row(ws_cat, i, [row.get(k, "") for k in cat_keys], alternate=(i % 2 == 1))
    _apply_column_format(ws_cat, col_index=4, fmt=_FMT_CURRENCY, start_row=2)
    _write_total_row(ws_cat, len(by_category_rows) + 2, [
        "TOTAL",
        sum(r["item_count"] for r in by_category_rows),
        sum(r["total_quantity"] for r in by_category_rows),
        grand_total["total_sales"],
    ])
    _auto_size_columns(ws_cat)

    wb.save(filepath)


def write_summary(rows, columns, sheet_title, filepath, grand_total_row=None):
    """
    Generic writer for summary/aggregated data (e.g. by-movie or by-employee views).

    Args:
        rows:           list of dicts, values in the same order as `columns`
        columns:        list of column-header strings matching the dict keys in order
        sheet_title:    name for the worksheet tab
        filepath:       destination path string
        grand_total_row: optional list of values to append as a grand-total row
                         (must have the same length as `columns`)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    _write_header(ws, columns)

    for i, row in enumerate(rows, start=2):
        # Rows are dicts; extract values in column-declaration order.
        data = [row.get(_col_key(col), "") for col in columns]
        _write_data_row(ws, i, data, alternate=(i % 2 == 1))

    if grand_total_row is not None:
        _write_total_row(ws, len(rows) + 2, grand_total_row)

    _auto_size_columns(ws)
    wb.save(filepath)


# ── Internal helpers ──────────────────────────────────────────────────────────


def _write_header(ws, columns):
    """Writes a bold, dark-navy header row with the given column labels."""
    header_font = Font(bold=True, color=_COLOUR_HEADER_FG, name="Calibri", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor=_COLOUR_HEADER_BG)

    for col_idx, label in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Freeze the header row so it stays visible when scrolling.
    ws.freeze_panes = "A2"


def _write_data_row(ws, row_num, values, alternate=False):
    """
    Writes a list of values as one data row.

    Applies alternating background shading if `alternate` is True.
    Each value is written with the correct cell type (date, number, or string).
    """
    fill = PatternFill(fill_type="solid", fgColor=_COLOUR_ROW_ALT) if alternate else None
    font = Font(name="Calibri", size=10)

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=_coerce_value(value))
        cell.font = font
        if fill:
            cell.fill = fill


def _write_total_row(ws, row_num, values):
    """
    Writes a grand-total row with bold white text on a navy background.
    """
    total_font = Font(bold=True, color=_COLOUR_TOTAL_FG, name="Calibri", size=11)
    total_fill = PatternFill(fill_type="solid", fgColor=_COLOUR_TOTAL_BG)

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=_coerce_value(value))
        cell.font  = total_font
        cell.fill  = total_fill


def _apply_column_format(ws, col_index, fmt, start_row=2, end_row=None):
    """
    Applies a number/date format string to all cells in a column from start_row onward.

    Args:
        col_index: 1-based column number
        fmt:       Excel format string (e.g. '#,##0.00' or 'YYYY-MM-DD')
        start_row: first row to format (default 2 to skip the header)
        end_row:   last row to format (default: ws.max_row)
    """
    last = end_row or ws.max_row
    for row_num in range(start_row, last + 1):
        ws.cell(row=row_num, column=col_index).number_format = fmt


def _auto_size_columns(ws):
    """
    Adjusts each column's width to fit its widest cell value.

    Caps maximum width at 50 characters to prevent absurdly wide columns
    when cell content is very long (e.g. long movie titles).
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                # date objects render as ~10 chars; everything else: use string length.
                if isinstance(cell.value, (datetime.date, datetime.datetime)):
                    cell_len = 10
                else:
                    cell_len = len(str(cell.value))
                max_len = max(max_len, cell_len)

        # Add padding and enforce minimum/maximum bounds.
        adjusted = min(max(max_len + 3, 10), 50)
        ws.column_dimensions[col_letter].width = adjusted


def _coerce_value(value):
    """
    Ensures a value is stored as the correct Python type in the xlsx cell.

    openpyxl writes Python date → Excel date, int/float → number, str → text.
    This function makes sure we never pass ambiguous types that openpyxl
    would serialise differently than expected.
    """
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        # Keep as datetime — openpyxl handles the Excel serial number conversion.
        return value
    if isinstance(value, datetime.date):
        # openpyxl handles date objects correctly.
        return value
    if isinstance(value, float):
        # Round floats to 2 decimal places for display consistency.
        return round(value, 2)
    return value


def _col_key(col_label):
    """
    Converts a human-readable column label into a lowercase snake_case dict key.

    Used by write_summary() to look up values from row dicts using column names.
    Example: "Box Gross ($)" → "box_gross_($)"  (close enough for simple matching)

    In practice, the caller is expected to pass rows whose keys align with the
    column names — this is a convenience helper, not a robust normaliser.
    """
    return col_label.lower().replace(" ", "_").replace("-", "_")
